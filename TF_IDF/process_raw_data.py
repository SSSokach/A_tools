from openpyxl import load_workbook
import pickle
from tqdm import tqdm
import json
import numpy as np
np.random.seed(1234)
def write_pkl(obj, filename):
    with open(filename, 'wb') as f:
        pickle.dump(obj, f)


def read_pkl(filename):
    with open(filename, 'rb') as f:
        return pickle.load(f)


def load_json(path):
    with open(path, 'r', encoding='utf-8') as f:
        data = json.load(f)
    return data


def write_json(path, data):
    with open(path, 'w', encoding='utf-8') as f:
        json.dump(data, f, ensure_ascii=False)

    
# # 加载工作簿
# wb2 = load_workbook('./jddc_dataset.xlsx')
# # 获取sheet页
# ws = wb2['Sheet1']
# # 打印sheet页的颜色属性值
# dataset=[]
# for row in tqdm(ws['1:10924']):
#     line=[]
#     for cell in row[:6]:
#         line.append(cell.value)
#     dataset.append(line)

# 转换数据集格式
# raw_dataset=dataset[1:]
# i=0
# dataset=[]
# while i<len(raw_dataset):
    
#     print('new session')
#     session=[]
#     while i<len(raw_dataset):
#         print(i)
#         raw_sentence=raw_dataset[i]
        
#         if raw_sentence[0] == "无":
#             i+=1
#             break
#         sentence={
#             'sens':raw_sentence[1].replace(' ',''),
#             'person':raw_sentence[2],
#             'intent':raw_sentence[3],
#             'emotion':raw_sentence[4],
#             'talktech':raw_sentence[5]
#         }
#         session.append(sentence)
#         i+=1
#     dataset.append(session)
# write_json('jddc_raw_data.json',dataset)

# 压缩intent类别，确定query没有话术
# dataset=load_json('jddc_raw_data.json')
# intent_to_abstract_map=load_json('intent_to_abstract_map.json')
# for session in dataset:
#     for utterance in session:
#         if utterance['person']=='A' and utterance['talktech']=="无":
#             utterance['talktech']='无话术'
#         if utterance['intent'] in intent_to_abstract_map:
#             utterance['intent']=intent_to_abstract_map[utterance['intent']]
# write_json('./jddc_sessions.json',dataset)

# 划分数据集
dataset=load_json('jddc_sessions_keyword.json')
shuffled_indices = np.random.permutation(len(dataset))
# 8:1:1
indices=shuffled_indices[:int(0.8*(len(dataset)))]
train_dataset=[dataset[i] for i in indices]
write_json('jddc_sessions_train.json',train_dataset)
print(len(train_dataset))

indices=shuffled_indices[int(0.8*(len(dataset))):int(0.9*(len(dataset)))]
valid_dataset=[dataset[i] for i in indices]
write_json('jddc_sessions_valid.json',valid_dataset)
print(len(valid_dataset))

indices=shuffled_indices[int(0.9*(len(dataset))):]
test_dataset=[dataset[i] for i in indices]
write_json('jddc_sessions_test.json',test_dataset)
print(len(test_dataset))

#统计标签类型
dataset=load_json('jddc_sessions.json')
dataset_utterances=[ i for x in dataset for i in x]
talktech_dic={"无":-1}
intent_dic={"无":-1}
for utterance in dataset_utterances:
    talktech=utterance['talktech']
    if talktech not in talktech_dic:
        talktech_dic[talktech]=len(talktech_dic)-1
    
    
    intent=utterance['intent']
    if intent not in intent_dic:
        intent_dic[intent]=len(intent_dic)-1
write_json('intent_jddc.json',intent_dic)
print(intent_dic)
write_json('talktech_jddc.json',talktech_dic)
talktech_dic=load_json('talktech_jddc.json')
print(talktech_dic)

# 统计标签分布
dataset=load_json('jddc_sessions.json')
dataset_utterances=[ i for x in dataset for i in x]
talktech_jddc=load_json('talktech_jddc.json')
talktech_dic={}
for i in talktech_jddc:
    talktech_dic[i]=0
intent_to_abstract_map=load_json('intent_to_abstract_map.json')
abstract_to_intent_map=load_json('abstract_to_intent_map.json')
abstract_intent_dic={"无":0}
for i in abstract_to_intent_map:
    abstract_intent_dic[i]=0


for line in dataset_utterances:
    talktech_dic[line['talktech']]+=1
    abstract_intent_dic[line['intent']]+=1
print(talktech_dic)
print(abstract_intent_dic)

